"""
Excel exporter for job application data.
Creates a comprehensive workbook with multiple sheets for tracking and analysis.
"""
from datetime import date, datetime
from pathlib import Path
from typing import Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database.database import get_session
from database.repositories import RepositoryFactory
from database.models import (
    Job, JobMatch, JobStatus, Application, ApplicationStatus,
    CandidateProfile, DailyStatistics, ScreeningQuestion, ApplicationError, Resume
)
from utils.logger import get_logger


logger = get_logger(__name__)


# Style definitions
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


class ExcelExporter:
    """Exports job application data to Excel workbook."""

    def __init__(self, output_path: str = "output/job_applications.xlsx"):
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb = Workbook()

    async def export(self) -> str:
        """Export all data to Excel and return the file path."""
        async with get_session() as session:
            repos = RepositoryFactory(session)

            # Get all data
            jobs = await repos.jobs.get_all(limit=10000)
            profile = await repos.candidates.get_profile()
            stats = await repos.statistics.get_or_create_today()

            # Build lookup maps
            matches = {m.job_id: m for m in await self._get_all_matches(session)}
            applications = {a.job_id: a for a in await self._get_all_applications(session)}
            resumes = {r.job_id: r for r in await self._get_all_resumes(session)}
            screening_questions = await self._get_all_screening_questions(session)
            errors = await self._get_all_errors(session)

            # Create sheets
            self._create_applications_sheet(jobs, matches, applications, resumes)
            self._create_job_analysis_sheet(jobs, matches)
            self._create_daily_statistics_sheet(stats)
            self._create_candidate_profile_sheet(profile)
            self._create_errors_sheet(errors)

            # Save
            self.wb.save(self.output_path)
            logger.info(f"Excel exported to {self.output_path}")
            return str(self.output_path)

    async def _get_all_matches(self, session) -> List[JobMatch]:
        from sqlalchemy import select
        result = await session.execute(select(JobMatch))
        return list(result.scalars().all())

    async def _get_all_applications(self, session) -> List[Application]:
        from sqlalchemy import select
        result = await session.execute(select(Application))
        return list(result.scalars().all())

    async def _get_all_resumes(self, session) -> List[Resume]:
        from sqlalchemy import select
        from database.models import Resume
        result = await session.execute(select(Resume))
        return list(result.scalars().all())

    async def _get_all_screening_questions(self, session) -> List[ScreeningQuestion]:
        from sqlalchemy import select
        result = await session.execute(select(ScreeningQuestion))
        return list(result.scalars().all())

    async def _get_all_errors(self, session) -> List[ApplicationError]:
        from sqlalchemy import select
        result = await session.execute(select(ApplicationError))
        return list(result.scalars().all())

    def _style_header_row(self, ws, row: int, num_cols: int):
        """Apply header styling to a row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

    def _style_data_cell(self, ws, row: int, col: int, value: any = None):
        """Apply data cell styling."""
        cell = ws.cell(row=row, column=col)
        if value is not None:
            cell.value = value
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        return cell

    def _auto_width(self, ws, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths."""
        for col in ws.columns:
            max_length = min_width
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), max_width))
            ws.column_dimensions[col_letter].width = max_length + 2

    def _write_row(self, ws, row: int, data: List[any], start_col: int = 1):
        """Write a row of data."""
        for i, value in enumerate(data):
            self._style_data_cell(ws, row, start_col + i, value)

    # ==================== SHEET 1: Applications ====================
    def _create_applications_sheet(
        self,
        jobs: List[Job],
        matches: dict,
        applications: dict,
        resumes: dict
    ):
        ws = self.wb.active
        ws.title = "Applications"

        headers = [
            "Job ID", "Date Found", "Date Applied", "Job Title", "Company",
            "Location", "Remote Type", "Salary", "Source", "Job URL",
            "Application URL", "Match Score", "Skills Match", "Experience Match",
            "Resume File", "Resume Version", "Application Status", "Confirmation", "Notes"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Write data
        row = 2
        for job in jobs:
            match = matches.get(job.id)
            app = applications.get(job.id)
            resume = resumes.get(job.id)

            # Format salary
            salary = ""
            if job.salary_min or job.salary_max:
                parts = []
                if job.salary_min:
                    parts.append(f"${job.salary_min:,}")
                if job.salary_max:
                    parts.append(f"${job.salary_max:,}")
                salary = " - ".join(parts)
                if job.currency:
                    salary += f" {job.currency}"

            # Format match score details
            skills_match = ""
            exp_match = ""
            if match:
                strong = len(match.strong_matches) if match.strong_matches else 0
                partial = len(match.partial_matches) if match.partial_matches else 0
                missing = len(match.missing_requirements) if match.missing_requirements else 0
                skills_match = f"{strong} strong, {partial} partial, {missing} missing"
                # Experience match - derive from match data
                exp_match = "Calculated in Job Analysis"

            data = [
                job.id,
                job.discovered_at.strftime("%Y-%m-%d") if job.discovered_at else "",
                app.applied_at.strftime("%Y-%m-%d") if app and app.applied_at else "",
                job.title,
                job.company,
                job.location,
                job.remote_type.value if job.remote_type else "",
                salary,
                job.source,
                job.canonical_url,
                app.application_url if app else "",
                match.match_score if match else "",
                skills_match,
                exp_match,
                resume.filename if resume else "",
                resume.version if resume else "",
                app.status.value if app else JobStatus.DISCOVERED.value,
                app.confirmation if app and app.confirmation else "",
                app.human_intervention_reason if app and app.human_intervention_reason else "",
            ]

            self._write_row(ws, row, data)

            # Color-code match score
            if match:
                score_cell = ws.cell(row=row, column=12)
                if match.match_score >= 85:
                    score_cell.fill = GREEN_FILL
                elif match.match_score >= 75:
                    score_cell.fill = YELLOW_FILL
                else:
                    score_cell.fill = RED_FILL

            # Color-code application status
            if app:
                status_cell = ws.cell(row=row, column=17)
                if app.status == ApplicationStatus.APPLIED:
                    status_cell.fill = GREEN_FILL
                elif app.status in (ApplicationStatus.NEEDS_HUMAN_INPUT, ApplicationStatus.APPLYING):
                    status_cell.fill = YELLOW_FILL
                elif app.status == ApplicationStatus.FAILED:
                    status_cell.fill = RED_FILL

            row += 1

        # Set column widths
        widths = [8, 12, 12, 30, 25, 20, 12, 18, 15, 50, 50, 10, 35, 35, 30, 12, 18, 50, 40]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

    # ==================== SHEET 2: Job Analysis ====================
    def _create_job_analysis_sheet(self, jobs: List[Job], matches: dict):
        ws = self.wb.create_sheet("Job Analysis")

        headers = [
            "Job ID", "Required Skills", "Preferred Skills", "Tools",
            "Education", "Experience", "Keywords", "Matched Skills",
            "Missing Skills", "Match Explanation"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        row = 2
        for job in jobs:
            match = matches.get(job.id)

            data = [
                job.id,
                ", ".join(job.skills) if job.skills else "",
                "",  # Preferred skills not stored separately
                ", ".join(job.tools) if job.tools else "",
                ", ".join(job.education) if job.education else "",
                job.requirements or "",
                ", ".join(job.skills + job.tools) if (job.skills or job.tools) else "",
                ", ".join([m.get("skill", "") if isinstance(m, dict) else str(m) for m in match.strong_matches]) if match and match.strong_matches else "",
                ", ".join([m if isinstance(m, str) else m.get("skill", "") for m in match.missing_requirements]) if match and match.missing_requirements else "",
                match.reasoning if match else "",
            ]

            self._write_row(ws, row, data)
            row += 1

        widths = [8, 40, 40, 30, 30, 50, 40, 40, 40, 60]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"

    # ==================== SHEET 3: Daily Statistics ====================
    def _create_daily_statistics_sheet(self, stats: DailyStatistics):
        ws = self.wb.create_sheet("Daily Statistics")

        headers = ["Date", "Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        metrics = [
            ("Jobs Found", stats.jobs_found),
            ("Duplicates Removed", stats.duplicates_removed),
            ("Jobs Qualified", stats.jobs_qualified),
            ("Resumes Created", stats.resumes_created),
            ("Resumes Validated", stats.resumes_validated),
            ("Applications Submitted", stats.applications_submitted),
            ("Applications Failed", stats.applications_failed),
            ("Human Intervention Required", stats.human_intervention_required),
            ("Average Match Score", f"{stats.average_match_score:.1f}%" if stats.average_match_score else "N/A"),
        ]

        row = 2
        today_str = stats.date.strftime("%Y-%m-%d") if stats.date else date.today().strftime("%Y-%m-%d")
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=today_str).border = THIN_BORDER
            ws.cell(row=row, column=2, value=metric).border = THIN_BORDER
            ws.cell(row=row, column=3, value=value).border = THIN_BORDER
            row += 1

        # Top companies
        if stats.top_companies:
            ws.cell(row=row, column=1, value="").border = THIN_BORDER
            ws.cell(row=row, column=2, value="Top Companies").border = THIN_BORDER
            ws.cell(row=row, column=3, value=", ".join(stats.top_companies)).border = THIN_BORDER
            row += 1

        # Top job titles
        if stats.top_job_titles:
            ws.cell(row=row, column=1, value="").border = THIN_BORDER
            ws.cell(row=row, column=2, value="Top Job Titles").border = THIN_BORDER
            ws.cell(row=row, column=3, value=", ".join(stats.top_job_titles)).border = THIN_BORDER
            row += 1

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 30

    # ==================== SHEET 4: Candidate Profile ====================
    def _create_candidate_profile_sheet(self, profile: Optional[CandidateProfile]):
        ws = self.wb.create_sheet("Candidate Profile")

        if not profile:
            ws.cell(row=1, column=1, value="No candidate profile found").font = Font(italic=True)
            return

        headers = ["Category", "Item", "Details"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        row = 2

        def add_row(category: str, item: str, details: str = ""):
            nonlocal row
            ws.cell(row=row, column=1, value=category).border = THIN_BORDER
            ws.cell(row=row, column=2, value=item).border = THIN_BORDER
            ws.cell(row=row, column=3, value=details).border = THIN_BORDER
            row += 1

        # Personal info
        add_row("Personal", "Name", profile.name)
        add_row("Personal", "Email", profile.email)
        add_row("Personal", "Phone", profile.phone or "")
        add_row("Personal", "Location", f"{profile.city}, {profile.province}" if profile.city else "")
        add_row("Personal", "Work Authorization", profile.work_authorization or "")
        add_row("Personal", "LinkedIn", profile.linkedin_url or "")
        add_row("Personal", "GitHub", profile.github_url or "")
        add_row("Personal", "Portfolio", profile.portfolio_url or "")

        # Skills
        for skill in profile.technical_skills:
            if isinstance(skill, dict):
                add_row("Technical Skill", skill.get("name", ""), skill.get("proficiency", ""))
            else:
                add_row("Technical Skill", str(skill), "")

        for skill in profile.tools:
            if isinstance(skill, dict):
                add_row("Tool", skill.get("name", ""), skill.get("proficiency", ""))
            else:
                add_row("Tool", str(skill), "")

        # Experience
        for emp in profile.employment_history:
            if isinstance(emp, dict):
                details = f"{emp.get('company', '')} | {emp.get('start_date', '')}-{emp.get('end_date', '')} | {emp.get('description', '')[:100]}"
                add_row("Employment", emp.get("title", ""), details)
            else:
                add_row("Employment", str(emp), "")

        # Education
        for edu in profile.education:
            if isinstance(edu, dict):
                details = f"{edu.get('institution', '')} | {edu.get('graduation_year', '')} | {edu.get('details', '')}"
                add_row("Education", edu.get("degree", ""), details)
            else:
                add_row("Education", str(edu), "")

        # Certifications
        for cert in profile.certifications:
            if isinstance(cert, dict):
                details = f"{cert.get('issuer', '')} | {cert.get('year', '')}"
                add_row("Certification", cert.get("name", ""), details)
            else:
                add_row("Certification", str(cert), "")

        # Preferences
        for title in profile.preferred_job_titles:
            add_row("Preferred Title", title, "")
        for loc in profile.preferred_locations:
            add_row("Preferred Location", loc, "")
        for remote in profile.remote_preferences:
            add_row("Remote Preference", remote, "")

        # Exclusions
        for title in profile.excluded_titles:
            add_row("Excluded Title", title, "")
        for ind in profile.excluded_industries:
            add_row("Excluded Industry", ind, "")
        for req in profile.excluded_requirements:
            add_row("Excluded Requirement", req, "")

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 60

    # ==================== SHEET 5: Application Errors ====================
    def _create_errors_sheet(self, errors: List[ApplicationError]):
        ws = self.wb.create_sheet("Application Errors")

        headers = [
            "Timestamp", "Job ID", "Source", "Error Type", "Error Message",
            "Current URL", "Resolution", "Resolved"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        row = 2
        for error in errors:
            data = [
                error.timestamp.strftime("%Y-%m-%d %H:%M:%S") if error.timestamp else "",
                error.application_id,  # This is application_id, not job_id
                error.source,
                error.error_type,
                error.error_message,
                error.current_url or "",
                error.resolution or "",
                "Yes" if error.resolved else "No",
            ]
            self._write_row(ws, row, data)

            # Color-code resolved
            resolved_cell = ws.cell(row=row, column=8)
            if error.resolved:
                resolved_cell.fill = GREEN_FILL
            else:
                resolved_cell.fill = RED_FILL

            row += 1

        widths = [20, 10, 15, 20, 50, 50, 30, 10]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"


async def export_to_excel(output_path: str = "output/job_applications.xlsx") -> str:
    """Convenience function to export data to Excel."""
    exporter = ExcelExporter(output_path)
    return await exporter.export()